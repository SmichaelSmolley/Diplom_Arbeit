import subprocess
import re
import os
from datetime import datetime, timedelta

from openpyxl import load_workbook

# =====================================================
# Einstellungen
# =====================================================

REPO = r"C:\Users\Michael Watzek\Documents\Diplom_Arbeit"

EXCEL_DATEI = os.path.join(REPO, "worktime.xlsx")

# =====================================================
# Excel öffnen
# =====================================================

wb = load_workbook(EXCEL_DATEI)
ws = wb["Begleitprotokoll"]

# =====================================================
# Erste freie Zeile suchen
# =====================================================

row = 5

while ws[f"A{row}"].value is not None:
    row += 1

print(f"Erste freie Zeile: {row}")

# =====================================================
# Git Log lesen
# =====================================================

result = subprocess.run(
    [
        "git",
        "-C",
        REPO,
        "log",
        "--reverse",
        "--pretty=format:%ad - %s",
        "--date=format:%d.%m.%Y %H:%M:%S"
    ],
    capture_output=True,
    text=True,
    encoding="utf-8"
)

eingetragen = 0

# =====================================================
# Commits verarbeiten
# =====================================================

for line in result.stdout.splitlines():

    try:
        date_str, message = line.split(" - ", 1)
    except ValueError:
        continue

    # Merge Commits ignorieren
    if "Merge branch" in message:
        continue

    # Zeitangabe suchen
    m = re.search(r"-\s*(\d+)\s*min", message, re.IGNORECASE)
    h = re.search(r"-\s*(\d+)\s*h", message, re.IGNORECASE)

    if m:
        minutes = int(m.group(1))
    elif h:
        minutes = int(h.group(1)) * 60
    else:
        print("Keine Zeitangabe:", message)
        continue

    dt = datetime.strptime(
        date_str,
        "%d.%m.%Y %H:%M:%S"
    )

    start = dt - timedelta(minutes=minutes)

    # Tätigkeit inklusive Zeitangabe übernehmen
    taetigkeit = message

    # =================================================
    # Formeln aus vorheriger Zeile übernehmen
    # =================================================

    for col in ["B", "F", "G"]:

        prev = ws[f"{col}{row-1}"].value

        if isinstance(prev, str) and prev.startswith("="):
            ws[f"{col}{row}"] = prev.replace(
                str(row - 1),
                str(row)
            )

    # =================================================
    # Daten eintragen
    # =================================================

    ws[f"A{row}"] = dt.date()
    ws[f"C{row}"] = start.time()
    ws[f"D{row}"] = dt.time()

    # Schule immer 0
    ws[f"E{row}"] = 0

    ws[f"H{row}"] = taetigkeit

    row += 1
    eingetragen += 1

# =====================================================
# Speichern
# =====================================================

wb.save(EXCEL_DATEI)

print()
print(f"{eingetragen} Commits eingetragen.")
print(f"Datei gespeichert:")
print(EXCEL_DATEI)